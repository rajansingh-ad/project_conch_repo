# -*- coding: utf-8 -*-

# =============================================================================
# FINAL ODI VERSION - NIKE FORECAST + BOM USAGE - ALL LOBS
# - Processes LOBs sequentially: Apparel -> Footwear -> Lion Brothers -> Accessories
# - Processes one or many source files sequentially inside every LOB
# - Skips an exact source output when the CSV filename already exists
# - ZIP metadata source_file_name is stored as ZIP_NAME/INNER_WORKBOOK_NAME
# - Lion Brothers uses the second date occurrence as file_creation_date
# - Forecast CSV locations remain unchanged; generated names use single underscores
# - BOM Usage CSVs are written below file_conversion_to_csv/bom_usage/<lob>
# - BOM Usage is appended exactly once as _BOM_Usage.csv
# =============================================================================

import os
import re
import sys
import csv
import time
import zipfile
import tempfile
from datetime import datetime, date
from collections import OrderedDict

from openpyxl import load_workbook


DEFAULT_INPUT_BASE = "/mnt/fss-rodn-iad-odi-fs/project_conch/nike"
DEFAULT_OUTPUT_BASE = "/mnt/fss-rodn-iad-odi-fs/project_conch/nike/file_conversion_to_csv"

INPUT_BASE = os.environ.get("NIKE_INPUT_BASE", DEFAULT_INPUT_BASE).strip()
OUTPUT_BASE = os.environ.get("NIKE_CSV_OUTPUT_BASE", DEFAULT_OUTPUT_BASE).strip()
BOM_OUTPUT_BASE = os.environ.get(
    "NIKE_BOM_CSV_OUTPUT_BASE",
    os.path.join(OUTPUT_BASE, "bom_usage")
).strip()

OVERWRITE_EXISTING = os.environ.get("OVERWRITE_EXISTING", "N").strip().upper() in ("Y", "YES", "TRUE", "1")
LOB_FILTER = os.environ.get("LOB_FILTER", "ALL").strip().lower()
MAX_FILES_PER_LOB = int(os.environ.get("MAX_FILES_PER_LOB", "0"))
ADD_METADATA_COLUMNS = os.environ.get("ADD_METADATA_COLUMNS", "Y").strip().upper() in ("Y", "YES", "TRUE", "1")
ENABLE_BOM_USAGE = os.environ.get("ENABLE_BOM_USAGE", "Y").strip().upper() in ("Y", "YES", "TRUE", "1")

HEADER_SCAN_ROWS = 80
MAX_COLS_TO_SCAN = 1500
MAX_CONSECUTIVE_BLANK_ROWS = 300
OUTPUT_BUFFER_SIZE = 1024 * 1024

EXCEL_EXTENSIONS = (".xlsx", ".xlsm")
OLD_XLS_EXTENSION = ".xls"
ZIP_EXTENSION = ".zip"

SKIP_DIRECTORY_NAMES = {
    "archive", "archived", "validation_reports",
    "file_conversion_to_csv", "__macosx", ".trash"
}

LOB_CONFIGS = OrderedDict([
    ("apparel", {
        "input_folder": "apparel",
        "output_folder": "apparel",
        "display_name": "Apparel",

        # FINAL BUSINESS RULE:
        # If an FM ZIP contains PFF workbooks, skip those PFF workbooks
        # successfully. Do not convert them and do not fail the full run.
        "sheet_keywords": [
            "ap production plan - fm",
            "production plan - fm",
            "ap production plan fm",
            "production plan fm",
            "finished material",
            "fm",
        ],
        "sheet_exclude_keywords": [
            "pff",
            "bom usage",
            "bom_usage",
            "pivot",
            "summary",
            "raw material",
            "plan run date",
        ],
        "zip_member_include_keywords": [
            "production plan fm",
            "production plan - fm",
            "finished material",
        ],
        "zip_member_exclude_keywords": [
            "pff",
            "plan run date",
            "run date",
            "readme",
            "summary",
            "pivot",
            "bom usage",
            "bom_usage",
            "raw material",
        ],
        "header_keywords": [
            "material",
            "factory",
            "season",
            "buy",
            "forecast",
            "planning",
            "quantity",
            "finished material",
            "supplier",
        ],
        "file_date_rule": "first_date",
        "bom_enabled": True,
    }),
    ("footwear", {
        "input_folder": "footwear",
        "output_folder": "footwear",
        "display_name": "Footwear",
        "sheet_keywords": ["fw production plan - fm", "production plan - fm", "finished material", "fm"],
        "sheet_exclude_keywords": ["bom usage", "bom_usage", "pivot", "summary", "raw material", "pff"],
        "header_keywords": ["material", "factory", "season", "buy", "forecast", "planning", "quantity", "finished material", "supplier"],
        "file_date_rule": "first_date",
        "bom_enabled": True,
    }),
    ("lion_brothers", {
        "input_folder": "lion_brothers",
        "output_folder": "lion_brothers",
        "display_name": "Lion Brothers",
        "sheet_keywords": ["ap production plan - fm", "production plan - fm", "finished material", "fm"],
        "sheet_exclude_keywords": ["bom usage", "bom_usage", "pivot", "summary", "raw material", "pff"],
        "header_keywords": ["material", "factory", "season", "buy", "forecast", "planning", "quantity", "finished material", "supplier"],
        # Business rule: Lion Brothers filenames can contain two business
        # dates. The second date is the file creation date required in ADW.
        "file_date_rule": "second_date",
        "bom_enabled": True,
    }),
    ("accessories", {
        "input_folder": "accessories",
        "output_folder": "accessories",
        "display_name": "Accessories",
        "sheet_keywords": ["page"],
        "sheet_exclude_keywords": ["pivot", "summary", "bom usage", "bom_usage"],
        "header_keywords": ["division", "segment", "brand", "style", "color", "planning", "product", "category", "season", "month", "year"],
        "file_date_rule": "first_date",
        "bom_enabled": False,
    }),
])


def build_dataset_config(lob_key, base_config, dataset_type):
    """Return an isolated config for FORECAST or BOM_USAGE conversion."""
    config = dict(base_config)
    config["dataset_type"] = dataset_type

    if dataset_type == "FORECAST":
        return config

    config["display_name"] = base_config["display_name"] + " BOM Usage"
    config["sheet_keywords"] = ["bom usage", "bom_usage"]
    config["sheet_exclude_keywords"] = ["pivot", "summary"]
    config["require_sheet_keyword_match"] = True
    config["header_keywords"] = [
        "planning material",
        "planning material code",
        "planning material name",
        "material",
        "factory location",
        "season",
        "fg size",
        "bom usage",
        "usage",
    ]
    config["required_in_zip"] = True

    # Apparel receives BOM Usage as a separate workbook inside the ZIP.
    # Footwear and Lion Brothers receive BOM Usage as a sheet inside the same
    # forecast workbook, so all workbooks remain eligible for sheet discovery.
    if lob_key == "apparel":
        config["zip_member_include_keywords"] = ["bom usage", "bom_usage"]
        config["zip_member_exclude_keywords"] = [
            "pff", "plan run date", "run date", "readme", "summary", "pivot"
        ]
    else:
        config["zip_member_include_keywords"] = []
        config["zip_member_exclude_keywords"] = ["readme", "summary", "pivot"]

    return config


def log(message):
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " | " + str(message), flush=True)


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = ("%.15f" % value).rstrip("0").rstrip(".")
        return "0" if text == "-0" else text
    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_text(value):
    return re.sub(r"\s+", " ", clean_text(value).lower()).strip()


def sanitize_name(value):
    base = os.path.splitext(os.path.basename(str(value)))[0]
    base = re.sub(r"[^A-Za-z0-9_.-]+", "_", base)
    return base.strip("_")


def normalize_generated_name(value):
    """Normalize separators introduced in a generated CSV filename."""
    name = sanitize_name(value)
    name = re.sub(r"_{2,}", "_", name)
    return name.strip("_.-")


def remove_trailing_bom_usage(value):
    """Remove one or more trailing BOM Usage suffixes before canonicalizing."""
    name = value
    bom_suffix = re.compile(
        r"(?:[_ .-]*BOM[_ .-]*Usage)+$",
        flags=re.IGNORECASE,
    )

    previous = None
    while name != previous:
        previous = name
        name = bom_suffix.sub("", name).rstrip("_.- ")

    return name


def make_unique_headers(values):
    headers = []
    seen = {}
    for index, value in enumerate(values):
        header = clean_text(value)
        if header == "":
            header = "UNNAMED_COL_" + str(index + 1)
        if header in seen:
            seen[header] += 1
            header = header + "_" + str(seen[header])
        else:
            seen[header] = 1
        headers.append(header)

    last_real_index = -1
    for index, header in enumerate(headers):
        if not header.startswith("UNNAMED_COL_"):
            last_real_index = index
    if last_real_index < 0:
        return []
    return headers[:last_real_index + 1]


def row_has_data(values):
    return any(clean_text(value) != "" for value in values)


def row_header_score(values, config):
    non_empty = [clean_text(value) for value in values if clean_text(value) != ""]
    if len(non_empty) < 3:
        return 0
    score = len(non_empty)
    joined = " ".join(value.lower() for value in non_empty)
    for keyword in config.get("header_keywords", []):
        if keyword.lower() in joined:
            score += 15
    score += len(set(normalize_text(value) for value in non_empty))
    return score


def detect_header_row(ws, config):
    max_col = min(ws.max_column or MAX_COLS_TO_SCAN, MAX_COLS_TO_SCAN)
    best_row = None
    best_score = -1
    best_values = []

    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, max_col=max_col, values_only=True),
        start=1
    ):
        values = list(row)
        score = row_header_score(values, config)
        if score > best_score:
            best_row = row_number
            best_score = score
            best_values = values

    headers = make_unique_headers(best_values)
    if best_row is None or best_score <= 0 or not headers:
        raise Exception("Could not detect a valid header row in sheet: " + ws.title)

    return {
        "header_row_number": best_row,
        "header_score": best_score,
        "headers": headers,
        "column_count": len(headers),
    }


def sheet_name_match_score(sheet_name, config):
    normalized = normalize_text(sheet_name)
    for excluded in config.get("sheet_exclude_keywords", []):
        if normalize_text(excluded) in normalized:
            return -100000

    score = 0
    for keyword in config.get("sheet_keywords", []):
        keyword_norm = normalize_text(keyword)
        if normalized == keyword_norm:
            score = max(score, 10000)
        elif keyword_norm in normalized:
            score = max(score, 5000)
    return score


def select_sheet_and_header(workbook, config):
    best = None

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        name_score = sheet_name_match_score(sheet_name, config)
        if name_score <= -100000:
            continue
        if config.get("require_sheet_keyword_match") and name_score <= 0:
            continue

        try:
            header_info = detect_header_row(ws, config)
        except Exception:
            continue

        current = {
            "sheet_name": sheet_name,
            "sheet": ws,
            "total_score": name_score + header_info["header_score"],
            "header_info": header_info,
        }

        if best is None or current["total_score"] > best["total_score"]:
            best = current

    if best is None:
        raise Exception(
            "No suitable " + config.get("dataset_type", "FORECAST")
            + " sheet found. Available sheets: " + str(workbook.sheetnames)
        )

    return best


MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}


def safe_date(year, month, day):
    try:
        return date(int(year), int(month), int(day))
    except Exception:
        return None


def date_candidates_from_filename(file_name):
    """
    Return dates in the same left-to-right order in which they occur in the
    filename. This is important for Lion Brothers: the second date occurrence
    is the file_creation_date even when it is earlier than the first date.
    """
    text = str(file_name).replace("\\", "/")
    positioned = []

    patterns = [
        (
            re.compile(
                r"((?:19|20)\d{2})[\s_\-]*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[\s_\-]*(\d{1,2})",
                re.IGNORECASE
            ),
            lambda match: safe_date(
                match.group(1),
                MONTH_MAP[match.group(2).lower()],
                match.group(3)
            )
        ),
        (
            re.compile(
                r"(\d{1,2})[\s_\-]*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[\s_\-]*((?:19|20)\d{2})",
                re.IGNORECASE
            ),
            lambda match: safe_date(
                match.group(3),
                MONTH_MAP[match.group(2).lower()],
                match.group(1)
            )
        ),
        (
            re.compile(
                r"(?<!\d)((?:19|20)\d{2})[\-_]?(\d{2})[\-_]?(\d{2})(?!\d)"
            ),
            lambda match: safe_date(
                match.group(1),
                match.group(2),
                match.group(3)
            )
        ),
        (
            re.compile(
                # Separators are mandatory here. With optional separators,
                # a Lion Brothers value such as Dec01_011958 could be
                # misread as 01-Jan-1958 by combining the day with HHMMSS.
                r"(?<!\d)(\d{2})[\-_](\d{2})[\-_]((?:19|20)\d{2})(?!\d)"
            ),
            lambda match: (
                safe_date(match.group(3), match.group(1), match.group(2))
                or safe_date(match.group(3), match.group(2), match.group(1))
            )
        ),
    ]

    for pattern, parser in patterns:
        for match in pattern.finditer(text):
            parsed = parser(match)
            if parsed:
                positioned.append((match.start(), match.end(), parsed))

    positioned.sort(key=lambda item: (item[0], item[1]))

    ordered_dates = []
    seen_occurrences = set()

    for start_pos, end_pos, parsed in positioned:
        occurrence_key = (start_pos, end_pos, parsed.isoformat())
        if occurrence_key in seen_occurrences:
            continue
        seen_occurrences.add(occurrence_key)
        ordered_dates.append(parsed)

    return ordered_dates

def derive_file_creation_date(source_name, physical_path, config):
    candidates = date_candidates_from_filename(source_name)
    if candidates:
        date_rule = config.get("file_date_rule", "first_date")
        if date_rule == "second_date":
            # Some historical Lion Brothers filenames contain only one valid
            # date. Use it as a compatibility fallback, but never use mtime
            # when a filename date is available.
            selected = candidates[1] if len(candidates) >= 2 else candidates[0]
        elif date_rule == "last_date":
            selected = candidates[-1]
        else:
            selected = candidates[0]
        return selected.strftime("%Y-%m-%d")
    return datetime.fromtimestamp(os.path.getmtime(physical_path)).strftime("%Y-%m-%d")


def build_output_headers(source_headers):
    headers = list(source_headers)
    normalized = [normalize_text(header).replace(" ", "_") for header in headers]

    if ADD_METADATA_COLUMNS:
        for metadata_header in ["source_file_name", "source_sheet_name", "file_creation_date", "etl_insert_ts"]:
            if metadata_header not in normalized:
                headers.append(metadata_header)
                normalized.append(metadata_header)

    return headers


def output_path_for_source(output_dir, source_name, zip_container_name="", dataset_type="FORECAST"):
    source_base = normalize_generated_name(source_name)
    zip_base = normalize_generated_name(zip_container_name) if zip_container_name else ""

    # A ZIP and its inner workbook are separate filename components. Join them
    # with one underscore only; never introduce a double underscore.
    if zip_base:
        output_name = zip_base + "_" + source_base
    else:
        output_name = source_base

    output_name = re.sub(r"_{2,}", "_", output_name).strip("_.-")

    if dataset_type == "BOM_USAGE":
        # Normalize legacy inputs such as:
        #   ...__BOM_Usage_BOM_USAGE
        # to exactly:
        #   ..._BOM_Usage
        output_name = remove_trailing_bom_usage(output_name)
        output_name = output_name + "_BOM_Usage"

    output_name += ".csv"
    return os.path.join(output_dir, output_name)


def convert_xlsx_to_csv(excel_path, source_name, output_file, config):
    workbook = load_workbook(excel_path, read_only=True, data_only=True)

    try:
        selected = select_sheet_and_header(workbook, config)
        ws = selected["sheet"]
        sheet_name = selected["sheet_name"]
        header_info = selected["header_info"]

        source_headers = header_info["headers"]
        output_headers = build_output_headers(source_headers)
        source_col_count = len(source_headers)
        data_start_row = header_info["header_row_number"] + 1

        file_creation_date = derive_file_creation_date(source_name, excel_path, config)
        etl_insert_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        if os.path.exists(output_file) and not OVERWRITE_EXISTING:
            log("SKIP_EXISTING_OUTPUT | " + output_file)
            return {
                "status": "SKIPPED", "rows_written": 0, "sheet_name": sheet_name,
                "header_row_number": header_info["header_row_number"], "output_file": output_file,
                "file_creation_date": file_creation_date,
                "dataset_type": config.get("dataset_type", "FORECAST")
            }

        temp_output = output_file + ".copying"
        if os.path.exists(temp_output):
            os.remove(temp_output)

        rows_written = 0
        blank_streak = 0
        source_norm = [normalize_text(header).replace(" ", "_") for header in source_headers]

        with open(temp_output, "w", encoding="utf-8", newline="", buffering=OUTPUT_BUFFER_SIZE) as handle:
            writer = csv.writer(handle, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(output_headers)

            for row in ws.iter_rows(min_row=data_start_row, max_col=source_col_count, values_only=True):
                source_values = list(row)

                if not row_has_data(source_values):
                    blank_streak += 1
                    if blank_streak >= MAX_CONSECUTIVE_BLANK_ROWS:
                        break
                    continue

                blank_streak = 0
                output_values = [clean_text(value) for value in source_values]

                if ADD_METADATA_COLUMNS:
                    if "source_file_name" not in source_norm:
                        output_values.append(source_name)
                    if "source_sheet_name" not in source_norm:
                        output_values.append(sheet_name)
                    if "file_creation_date" not in source_norm:
                        output_values.append(file_creation_date)
                    if "etl_insert_ts" not in source_norm:
                        output_values.append(etl_insert_ts)

                writer.writerow(output_values)
                rows_written += 1

        os.replace(temp_output, output_file)

        log(
            "CONVERTED | LOB=" + config["display_name"]
            + " | Dataset=" + config.get("dataset_type", "FORECAST")
            + " | Source=" + source_name
            + " | Sheet=" + sheet_name
            + " | HeaderRow=" + str(header_info["header_row_number"])
            + " | Rows=" + str(rows_written)
            + " | Output=" + output_file
        )

        return {
            "status": "SUCCESS", "rows_written": rows_written, "sheet_name": sheet_name,
            "header_row_number": header_info["header_row_number"], "output_file": output_file,
            "file_creation_date": file_creation_date,
            "dataset_type": config.get("dataset_type", "FORECAST")
        }

    finally:
        workbook.close()


def convert_old_xls_to_csv(excel_path, source_name, output_file, config):
    try:
        import pandas as pd
    except ImportError:
        raise Exception("pandas is required for old .xls files: " + excel_path)

    xl = pd.ExcelFile(excel_path, engine="xlrd")
    best_sheet = None
    best_score = -1
    best_header_index = None

    for sheet_name in xl.sheet_names:
        name_score = sheet_name_match_score(sheet_name, config)
        if name_score <= -100000:
            continue
        if config.get("require_sheet_keyword_match") and name_score <= 0:
            continue

        top_df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, engine="xlrd", dtype=object, nrows=HEADER_SCAN_ROWS)

        for index in range(len(top_df)):
            score = name_score + row_header_score(top_df.iloc[index].tolist(), config)
            if score > best_score:
                best_score = score
                best_sheet = sheet_name
                best_header_index = index

    if best_sheet is None:
        raise Exception("No suitable sheet found in .xls file: " + excel_path)

    df = pd.read_excel(excel_path, sheet_name=best_sheet, header=best_header_index, engine="xlrd", dtype=object)
    df = df.dropna(how="all")
    df.columns = make_unique_headers(list(df.columns))

    normalized = [normalize_text(x).replace(" ", "_") for x in df.columns]
    if ADD_METADATA_COLUMNS:
        if "source_file_name" not in normalized:
            df["source_file_name"] = source_name
        if "source_sheet_name" not in normalized:
            df["source_sheet_name"] = best_sheet
        if "file_creation_date" not in normalized:
            df["file_creation_date"] = derive_file_creation_date(source_name, excel_path, config)
        if "etl_insert_ts" not in normalized:
            df["etl_insert_ts"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(output_file) and not OVERWRITE_EXISTING:
        log("SKIP_EXISTING_OUTPUT | " + output_file)
        return {
            "status": "SKIPPED", "rows_written": 0, "sheet_name": best_sheet,
            "header_row_number": best_header_index + 1, "output_file": output_file,
            "file_creation_date": derive_file_creation_date(source_name, excel_path, config),
            "dataset_type": config.get("dataset_type", "FORECAST")
        }

    temp_output = output_file + ".copying"
    df.to_csv(temp_output, index=False, encoding="utf-8")
    os.replace(temp_output, output_file)

    return {
        "status": "SUCCESS", "rows_written": len(df), "sheet_name": best_sheet,
        "header_row_number": best_header_index + 1, "output_file": output_file,
        "file_creation_date": derive_file_creation_date(source_name, excel_path, config),
        "dataset_type": config.get("dataset_type", "FORECAST")
    }


def process_excel_file(excel_path, source_name, output_dir, config, zip_container_name=""):
    dataset_type = config.get("dataset_type", "FORECAST")
    output_file = output_path_for_source(
        output_dir, source_name, zip_container_name, dataset_type=dataset_type
    )

    # Check before opening the workbook. If the exact converted output already
    # exists, skip it immediately and continue with the next file/LOB.
    if os.path.exists(output_file) and not OVERWRITE_EXISTING:
        log(
            "SKIP_EXISTING_OUTPUT"
            + " | LOB=" + config["display_name"]
            + " | Source=" + source_name
            + " | Output=" + output_file
        )
        return {
            "status": "SKIPPED",
            "rows_written": 0,
            "sheet_name": "",
            "header_row_number": "",
            "output_file": output_file,
            "file_creation_date": derive_file_creation_date(source_name, excel_path, config),
            "dataset_type": dataset_type
        }

    if excel_path.lower().endswith(EXCEL_EXTENSIONS):
        return convert_xlsx_to_csv(excel_path, source_name, output_file, config)

    if excel_path.lower().endswith(OLD_XLS_EXTENSION):
        return convert_old_xls_to_csv(excel_path, source_name, output_file, config)

    raise Exception("Unsupported Excel extension: " + excel_path)



def zip_member_is_eligible(member_name, config):
    """Return True only when an inner workbook is eligible for conversion."""
    normalized_name = normalize_text(os.path.basename(member_name))

    for keyword in config.get("zip_member_exclude_keywords", []):
        if normalize_text(keyword) in normalized_name:
            return False

    include_keywords = config.get("zip_member_include_keywords", [])
    if not include_keywords:
        return True

    return any(
        normalize_text(keyword) in normalized_name
        for keyword in include_keywords
    )

def process_zip_file(zip_path, output_dir, config, temp_root):
    results = []
    zip_name = os.path.basename(zip_path)
    extract_dir = os.path.join(temp_root, sanitize_name(zip_name))
    os.makedirs(extract_dir, exist_ok=True)

    with zipfile.ZipFile(zip_path, "r") as archive:
        excel_members = [
            member for member in archive.namelist()
            if not member.endswith("/")
            and "__macosx" not in member.lower()
            and not os.path.basename(member).startswith("~$")
            and (
                member.lower().endswith(EXCEL_EXTENSIONS)
                or member.lower().endswith(OLD_XLS_EXTENSION)
            )
        ]

        if not excel_members:
            raise Exception("No Excel files found inside ZIP: " + zip_path)

        eligible_members = []

        for member in sorted(excel_members):
            inner_name = os.path.basename(member)
            tracked_source_name = zip_name + "/" + inner_name

            if not zip_member_is_eligible(member, config):
                log(
                    "SKIP_INELIGIBLE_INNER_FILE"
                    + " | LOB=" + config["display_name"]
                    + " | ZIP=" + zip_name
                    + " | InnerFile=" + inner_name
                    + " | Reason=Business rule/helper workbook"
                )
                results.append({
                    "status": "SKIPPED",
                    "rows_written": 0,
                    "sheet_name": "",
                    "header_row_number": "",
                    "output_file": "",
                    "error_message": "",
                    "source_file_name": tracked_source_name,
                    "dataset_type": config.get("dataset_type", "FORECAST"),
                    "file_creation_date": derive_file_creation_date(
                        tracked_source_name, zip_path, config
                    ),
                })
                continue

            eligible_members.append(member)

        # Forecast-only PFF/helper ZIPs remain successful skips. A missing BOM
        # workbook is a failure because cleanup must not occur without BOM.
        if not eligible_members:
            if config.get("required_in_zip"):
                message = (
                    "No eligible " + config.get("dataset_type", "BOM_USAGE")
                    + " workbook found inside ZIP: " + zip_name
                )
                log("FAILED_REQUIRED_ZIP_CONTENT | " + message)
                results.append({
                    "status": "FAILED",
                    "rows_written": 0,
                    "sheet_name": "",
                    "header_row_number": "",
                    "output_file": "",
                    "error_message": message,
                    "source_file_name": zip_name,
                    "dataset_type": config.get("dataset_type", "BOM_USAGE"),
                    "file_creation_date": derive_file_creation_date(
                        zip_name, zip_path, config
                    ),
                })
                return results

            log(
                "SKIP_ZIP_NO_ELIGIBLE_FORECAST"
                + " | LOB=" + config["display_name"]
                + " | ZIP=" + zip_name
            )
            return results

        for member in eligible_members:
            inner_name = os.path.basename(member)
            tracked_source_name = zip_name + "/" + inner_name

            expected_output = output_path_for_source(
                output_dir,
                tracked_source_name,
                zip_container_name=zip_name,
                dataset_type=config.get("dataset_type", "FORECAST")
            )

            if os.path.exists(expected_output) and not OVERWRITE_EXISTING:
                log(
                    "SKIP_EXISTING_OUTPUT"
                    + " | LOB=" + config["display_name"]
                    + " | Source=" + tracked_source_name
                    + " | Output=" + expected_output
                )
                results.append({
                    "status": "SKIPPED",
                    "rows_written": 0,
                    "sheet_name": "",
                    "header_row_number": "",
                    "output_file": expected_output,
                    "error_message": "",
                    "source_file_name": tracked_source_name,
                    "dataset_type": config.get("dataset_type", "FORECAST"),
                    "file_creation_date": derive_file_creation_date(
                        tracked_source_name, zip_path, config
                    ),
                })
                continue

            extracted_path = archive.extract(member, extract_dir)

            try:
                result = process_excel_file(
                    extracted_path,
                    tracked_source_name,
                    output_dir,
                    config,
                    zip_container_name=zip_name
                )
                result["source_file_name"] = tracked_source_name
                results.append(result)

            except Exception as exc:
                log(
                    "FAILED_INNER_FILE"
                    + " | ZIP=" + zip_name
                    + " | InnerFile=" + inner_name
                    + " | Error=" + str(exc)
                )
                results.append({
                    "status": "FAILED",
                    "rows_written": 0,
                    "sheet_name": "",
                    "header_row_number": "",
                    "output_file": expected_output,
                    "error_message": str(exc),
                    "source_file_name": tracked_source_name,
                    "dataset_type": config.get("dataset_type", "FORECAST"),
                    "file_creation_date": derive_file_creation_date(
                        tracked_source_name, zip_path, config
                    ),
                })

    return results

def list_source_files(input_dir):
    files = []

    for root, dirs, names in os.walk(input_dir):
        dirs[:] = [directory for directory in dirs if directory.strip().lower() not in SKIP_DIRECTORY_NAMES]

        for name in names:
            lower = name.lower()
            if lower.startswith("~$"):
                continue

            full_path = os.path.join(root, name)

            if lower.endswith(EXCEL_EXTENSIONS) or lower.endswith(OLD_XLS_EXTENSION) or lower.endswith(ZIP_EXTENSION):
                files.append(full_path)

    files = sorted(files, key=lambda path: (os.path.getmtime(path), path.lower()))

    if MAX_FILES_PER_LOB > 0:
        files = files[:MAX_FILES_PER_LOB]

    return files


def process_lob(lob_key, config, summary_rows):
    input_dir = os.path.join(INPUT_BASE, config["input_folder"])
    forecast_output_dir = os.path.join(OUTPUT_BASE, config["output_folder"])
    bom_output_dir = os.path.join(BOM_OUTPUT_BASE, config["output_folder"])

    log("============================================================")
    log("LOB: " + config["display_name"])
    log("Input: " + input_dir)
    log("Forecast output: " + forecast_output_dir)
    if ENABLE_BOM_USAGE and config.get("bom_enabled"):
        log("BOM Usage output: " + bom_output_dir)

    if not os.path.exists(input_dir):
        log("SKIP_MISSING_INPUT_FOLDER | " + input_dir)
        return

    os.makedirs(forecast_output_dir, exist_ok=True)
    if ENABLE_BOM_USAGE and config.get("bom_enabled"):
        os.makedirs(bom_output_dir, exist_ok=True)

    dataset_jobs = [
        (
            "FORECAST",
            forecast_output_dir,
            build_dataset_config(lob_key, config, "FORECAST")
        )
    ]
    if ENABLE_BOM_USAGE and config.get("bom_enabled"):
        dataset_jobs.append(
            (
                "BOM_USAGE",
                bom_output_dir,
                build_dataset_config(lob_key, config, "BOM_USAGE")
            )
        )

    files = list_source_files(input_dir)
    log("Source files discovered: " + str(len(files)))

    with tempfile.TemporaryDirectory(prefix="nike_" + lob_key + "_excel_to_csv_") as temp_root:
        for file_path in files:
            source_name = os.path.basename(file_path)
            for dataset_type, output_dir, dataset_config in dataset_jobs:
                start_time = time.time()

                try:
                    if file_path.lower().endswith(ZIP_EXTENSION):
                        results = process_zip_file(
                            file_path, output_dir, dataset_config, temp_root
                        )
                        for result in results:
                            summary_rows.append({
                                "lob": config["display_name"],
                                "dataset_type": dataset_type,
                                "source_file_name": result.get(
                                    "source_file_name", source_name
                                ),
                                "file_creation_date": result.get(
                                    "file_creation_date", ""
                                ),
                                "status": result.get("status", ""),
                                "selected_sheet": result.get("sheet_name", ""),
                                "header_row_number": result.get("header_row_number", ""),
                                "rows_written": result.get("rows_written", 0),
                                "output_file": result.get("output_file", ""),
                                "elapsed_seconds": round(time.time() - start_time, 2),
                                "error_message": result.get("error_message", "")
                            })
                    else:
                        result = process_excel_file(
                            file_path, source_name, output_dir, dataset_config
                        )
                        summary_rows.append({
                            "lob": config["display_name"],
                            "dataset_type": dataset_type,
                            "source_file_name": source_name,
                            "file_creation_date": result.get("file_creation_date", ""),
                            "status": result.get("status", ""),
                            "selected_sheet": result.get("sheet_name", ""),
                            "header_row_number": result.get("header_row_number", ""),
                            "rows_written": result.get("rows_written", 0),
                            "output_file": result.get("output_file", ""),
                            "elapsed_seconds": round(time.time() - start_time, 2),
                            "error_message": ""
                        })

                except Exception as exc:
                    log(
                        "FAILED_FILE | LOB=" + config["display_name"]
                        + " | Dataset=" + dataset_type
                        + " | Source=" + source_name
                        + " | Error=" + str(exc)
                    )
                    summary_rows.append({
                        "lob": config["display_name"],
                        "dataset_type": dataset_type,
                        "source_file_name": source_name,
                        "file_creation_date": derive_file_creation_date(
                            source_name, file_path, dataset_config
                        ),
                        "status": "FAILED",
                        "selected_sheet": "",
                        "header_row_number": "",
                        "rows_written": 0,
                        "output_file": "",
                        "elapsed_seconds": round(time.time() - start_time, 2),
                        "error_message": str(exc)
                    })


def write_summary(summary_rows):
    log_dir = os.environ.get("LOG_DIR", os.path.join(INPUT_BASE, "logs")).strip()
    os.makedirs(log_dir, exist_ok=True)

    summary_file = os.path.join(
        log_dir,
        "nike_excel_to_csv_summary_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".csv"
    )

    fields = [
        "lob", "dataset_type", "source_file_name", "file_creation_date",
        "status", "selected_sheet",
        "header_row_number", "rows_written", "output_file",
        "elapsed_seconds", "error_message"
    ]

    with open(summary_file, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in summary_rows:
            writer.writerow(row)

    log("Conversion summary: " + summary_file)
    return summary_file


def main():
    log("============================================================")
    log("Nike Excel/ZIP to CSV Conversion Started")
    log("INPUT_BASE: " + INPUT_BASE)
    log("OUTPUT_BASE: " + OUTPUT_BASE)
    log("BOM_OUTPUT_BASE: " + BOM_OUTPUT_BASE)
    log("LOB_FILTER: " + LOB_FILTER)
    log("OVERWRITE_EXISTING: " + str(OVERWRITE_EXISTING))
    log("ADD_METADATA_COLUMNS: " + str(ADD_METADATA_COLUMNS))
    log("ENABLE_BOM_USAGE: " + str(ENABLE_BOM_USAGE))
    log("MAX_FILES_PER_LOB: " + str(MAX_FILES_PER_LOB))
    log("============================================================")

    if not os.path.exists(INPUT_BASE):
        raise Exception("Nike input base does not exist: " + INPUT_BASE)

    summary_rows = []
    start_time = time.time()

    for lob_key, config in LOB_CONFIGS.items():
        if LOB_FILTER not in ("", "all") and LOB_FILTER != lob_key:
            continue
        process_lob(lob_key, config, summary_rows)

    summary_file = write_summary(summary_rows)

    success_count = len([row for row in summary_rows if row["status"] == "SUCCESS"])
    skipped_count = len([row for row in summary_rows if row["status"] == "SKIPPED"])
    failed_count = len([row for row in summary_rows if row["status"] == "FAILED"])

    log("============================================================")
    log("Nike Excel/ZIP to CSV Conversion Completed")
    log("SUCCESS: " + str(success_count))
    log("SKIPPED: " + str(skipped_count))
    log("FAILED: " + str(failed_count))
    log("Elapsed seconds: " + str(round(time.time() - start_time, 2)))
    log("Summary file: " + summary_file)
    log("============================================================")

    if failed_count > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
